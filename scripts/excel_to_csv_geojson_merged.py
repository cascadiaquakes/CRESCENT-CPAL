# -*- coding: utf-8 -*-
import pandas as pd
import json
import re
import sys
import os
from openpyxl import load_workbook  
import numpy as np  
from datetime import datetime, date

NUM_HEADER_ROWS = 1


def clean_column_name(col_name):
    """Removes special characters like degree symbols and trims spaces."""
    col_name = re.sub(r"\[.*?\]", "", col_name)
    col_name = col_name.replace("_", " ").strip()
    return col_name


def process_excel_to_csv_geojson(input_file, num_header_rows=NUM_HEADER_ROWS):
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_csv = f"{base_name}.csv"
    output_geojson = f"{base_name}.geojson"

    # Step 1: Load Excel with pandas to get headers
    df = pd.read_excel(input_file, header=None, dtype=str)

    # Step 2: Process the headers (no change here)
    header_rows = df.iloc[:num_header_rows]
    header_rows = header_rows.ffill(axis=0).ffill(axis=1)

    def combine_headers(col_values):
        clean_vals = (
            col_values.dropna()
            .astype(str)
            .str.replace(r"[\r\n]+", " ", regex=True)
            .str.strip()
        )
        result = []
        for val in clean_vals:
            if not result or val != result[-1]:
                result.append(val)
        return "_".join(result)

    df.columns = header_rows.apply(combine_headers, axis=0)
    df.columns = [clean_column_name(col) for col in df.columns]

    # Step 3: Load body rows separately with openpyxl
    wb = load_workbook(filename=input_file, data_only=True)
    ws = wb.active

    # Precompute merged cells lookup
    merged_lookup = {}
    for merged_range in ws.merged_cells.ranges:
        tl_row = merged_range.min_row
        tl_col = merged_range.min_col
        tl_value = ws.cell(row=tl_row, column=tl_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = tl_value

    real_column_count = len(df.columns)

    body_rows = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=num_header_rows + 1), start=num_header_rows + 1
    ):
        row_data = []

        # Only process as many columns as we have in the header
        for col_idx, cell in enumerate(row[:real_column_count], start=1):
            value = merged_lookup.get((row_idx, col_idx), cell.value)
            row_data.append(np.nan if value is None else value)

        # 🔥 Skip completely empty rows (all values are NaN)
        if all(
            pd.isna(val) or (isinstance(val, str) and val.strip() == "")
            for val in row_data
        ):
            continue

        body_rows.append(row_data)

    # Now create the DataFrame
    body_df = pd.DataFrame(body_rows, columns=df.columns)

    # Step 5: Continue normal processing
    for col in body_df.columns:
        try:
            body_df[col] = pd.to_numeric(body_df[col])
        except (ValueError, TypeError):
            pass  # Keep as string if cannot convert

    # Save to CSV
    body_df.to_csv(output_csv, index=False)
    print(f"✅ Processed '{input_file}' → '{output_csv}'")

    # Step 6: Optional - create GeoJSON if lat/lon exist
    lat_col = next(
        (col for col in body_df.columns if re.search(r"latitude", col, re.IGNORECASE)),
        None,
    )
    lon_col = next(
        (col for col in body_df.columns if re.search(r"longitude", col, re.IGNORECASE)),
        None,
    )

    if lat_col and lon_col:
        print(
            f"🌍 Latitude column detected: '{lat_col}', Longitude column detected: '{lon_col}'"
        )
        geojson_features = []
        for _, row in body_df.iterrows():
            try:
                latitude = float(row[lat_col])
                longitude = float(row[lon_col])
                properties = (
                    row.to_dict()
                )  # We want Latitude and Longitude included row.drop([lat_col, lon_col]).to_dict()

                # Clean properties: replace NaN with empty strings and format keys
                properties = {}
                for (
                    key,
                    value,
                ) in (
                    row.items()
                ):  # See above note row.drop([lat_col, lon_col]).items():
                    clean_key = clean_column_name(key)
                    if pd.isna(value):
                        properties[clean_key] = ""
                    elif isinstance(value, (datetime, date)):
                        properties[clean_key] = value.strftime(
                            "%Y-%m-%d"
                        )  # Only date part
                    else:
                        properties[clean_key] = value

                geojson_features.append(
                    {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [longitude, latitude],
                        },
                        "properties": properties,
                    }
                )
            except ValueError:
                print(f"⚠️ Skipping invalid coordinates: {row[lat_col]}, {row[lon_col]}")

        geojson_data = {"type": "FeatureCollection", "features": geojson_features}
        with open(output_geojson, "w") as geojson_file:
            json.dump(geojson_data, geojson_file, indent=4)

        print(f"✅ GeoJSON file created: '{output_geojson}'")
    else:
        print("⚠️ No Latitude/Longitude columns found. Skipping GeoJSON output.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❌ Error: Please provide an Excel file as an argument.")
        print("Usage: python script.py <filename.xlsx>")
        sys.exit(1)

    input_filename = sys.argv[1]

    if not os.path.exists(input_filename):
        print(f"❌ Error: File '{input_filename}' not found.")
        sys.exit(1)

    process_excel_to_csv_geojson(input_filename)
